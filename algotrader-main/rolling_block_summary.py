"""
Rolling Block Program — Division-wise Summary Table Generator
============================================================
Divisions : BSP (Bilaspur), R (Raipur), NAG (Nagpur)
Block Type: CORRIDOR only
Output    : Dept-wise pivot — Total Hours Demanded vs Provided
            (split by Approved / Permitted)

USAGE
-----
1. Save your Excel/CSV as:  rolling_block_data.xlsx  (or .csv)
   in the same folder as this script.
2. Adjust the COLUMN_MAP below to match your actual column headers.
3. Run:  python rolling_block_summary.py
"""

import sys
import os
import pandas as pd

# ─── CONFIG ──────────────────────────────────────────────────────────────────
DATA_FILE = "rolling_block_data.xlsx"   # or .csv
SHEET     = 0                           # 0 = first sheet; change if needed

# Map your actual column names here  →  internal alias
COLUMN_MAP = {
    # YOUR COLUMN NAME       : INTERNAL NAME
    "Division"               : "division",
    "Disconnection ID"       : "disc_id",
    "Block Type"             : "block_type",
    "Department"             : "department",
    "Approved"               : "approved",     # e.g. "YES"/"NO" or 1/0
    "Permitted"              : "permitted",    # e.g. "YES"/"NO" or 1/0
    "Total Hours Demanded"   : "hrs_demanded",
    "Total Hours Provided"   : "hrs_provided",
}

DIVISION_CODES = {
    "BSP": "BSP",
    "R"  : "R",
    "NAG": "NAG",
    "NAg": "NAG",   # handle case variation
}

BLOCK_TYPE_CORRIDOR = "CORRIDOR"   # exact string that means corridor

# ─────────────────────────────────────────────────────────────────────────────

def load_data(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        df = pd.read_excel(path, sheet_name=SHEET, dtype=str)
    elif ext == ".csv":
        df = pd.read_csv(path, dtype=str)
    else:
        sys.exit(f"Unsupported file type: {ext}")
    df.columns = df.columns.str.strip()
    return df


def remap_columns(df):
    missing = [c for c in COLUMN_MAP if c not in df.columns]
    if missing:
        print("\n[WARNING] These mapped columns were NOT found in the file:")
        for c in missing:
            print(f"  '{c}'")
        print("\nAvailable columns:")
        for c in df.columns:
            print(f"  '{c}'")
        sys.exit("\nFix COLUMN_MAP in the script and retry.")
    return df.rename(columns=COLUMN_MAP)


def normalise(df):
    for col in ["division", "disc_id", "block_type", "department", "approved", "permitted"]:
        if col in df.columns:
            df[col] = df[col].fillna("").str.strip().str.upper()

    # Normalise numeric columns
    for col in ["hrs_demanded", "hrs_provided"]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Normalise division to standard codes
    div_upper = {k.upper(): v for k, v in DIVISION_CODES.items()}
    df["division"] = df["division"].map(div_upper).fillna(df["division"])

    return df


def build_summary(df):
    # ── Filter: Corridor blocks only ─────────────────────────────────────────
    df_corr = df[df["block_type"].str.contains(BLOCK_TYPE_CORRIDOR, case=False, na=False)]

    # ── Filter: disconnection ID starts with division code ───────────────────
    mask = df_corr.apply(
        lambda r: r["disc_id"].startswith(r["division"]), axis=1
    )
    df_corr = df_corr[mask]

    # ── Approved / Permitted flags ────────────────────────────────────────────
    # Treat "YES", "Y", "1" as True; everything else as False
    yes_vals = {"YES", "Y", "1", "TRUE", "APPROVED", "PERMITTED"}
    df_corr["is_approved"]  = df_corr["approved"].isin(yes_vals)
    df_corr["is_permitted"] = df_corr["permitted"].isin(yes_vals)

    target_divs = ["BSP", "R", "NAG"]
    results = []

    for div in target_divs:
        df_div = df_corr[df_corr["division"] == div]

        # Department-wise breakdown
        dept_grp = df_div.groupby("department", sort=True).agg(
            Total_Blocks        = ("disc_id",      "count"),
            Approved_Blocks     = ("is_approved",  "sum"),
            Permitted_Blocks    = ("is_permitted", "sum"),
            Hrs_Demanded        = ("hrs_demanded", "sum"),
            Hrs_Provided        = ("hrs_provided", "sum"),
        ).reset_index()
        dept_grp.insert(0, "Division", div)

        results.append(dept_grp)

        # Division total row
        total_row = pd.DataFrame([{
            "Division"        : div,
            "department"      : "** TOTAL **",
            "Total_Blocks"    : dept_grp["Total_Blocks"].sum(),
            "Approved_Blocks" : dept_grp["Approved_Blocks"].sum(),
            "Permitted_Blocks": dept_grp["Permitted_Blocks"].sum(),
            "Hrs_Demanded"    : dept_grp["Hrs_Demanded"].sum(),
            "Hrs_Provided"    : dept_grp["Hrs_Provided"].sum(),
        }])
        results.append(total_row)

    summary = pd.concat(results, ignore_index=True)
    summary.rename(columns={"department": "Department"}, inplace=True)

    # Utilisation %
    summary["Utilisation_%"] = (
        summary["Hrs_Provided"] / summary["Hrs_Demanded"].replace(0, pd.NA) * 100
    ).round(2)

    return summary


def print_table(df):
    pd.set_option("display.max_rows",    500)
    pd.set_option("display.max_columns", 20)
    pd.set_option("display.width",       140)
    pd.set_option("display.float_format", "{:,.2f}".format)
    print("\n" + "="*120)
    print(" ROLLING BLOCK PROGRAM — DIVISION-WISE SUMMARY  (CORRIDOR BLOCKS)")
    print("="*120)
    print(df.to_string(index=False))
    print("="*120 + "\n")


def export_excel(df, out_path="rolling_block_summary_output.xlsx"):
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Summary")
        ws = writer.sheets["Summary"]

        # Bold the TOTAL rows
        from openpyxl.styles import Font, PatternFill, Alignment
        yellow = PatternFill("solid", fgColor="FFFF99")
        blue   = PatternFill("solid", fgColor="BDD7EE")

        div_colors = {"BSP": "C6EFCE", "R": "FFEB9C", "NAG": "FFC7CE"}

        current_div = None
        for row in ws.iter_rows(min_row=2):
            div_val  = row[0].value
            dept_val = row[1].value
            if div_val != current_div:
                current_div = div_val
            fill_color = div_colors.get(str(div_val), "FFFFFF")
            if str(dept_val).startswith("**"):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor=fill_color)
                    cell.alignment = Alignment(horizontal="right")

        # Auto-width columns
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    print(f"[✓] Excel output saved → {out_path}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if not os.path.exists(DATA_FILE):
        print(f"\n[ERROR] Data file '{DATA_FILE}' not found.")
        print("  ► Save your spreadsheet as 'rolling_block_data.xlsx' (or .csv)")
        print("    in the same folder as this script, then re-run.\n")
        sys.exit(1)

    print(f"[...] Loading '{DATA_FILE}' ...")
    df = load_data(DATA_FILE)
    print(f"      {len(df):,} rows × {len(df.columns)} columns loaded.")

    df = remap_columns(df)
    df = normalise(df)

    summary = build_summary(df)
    print_table(summary)
    export_excel(summary)
